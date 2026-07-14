"""
Exportador a Excel que calca el formato del Tablero manual (TABLERO_ING):
encabezados agrupados fusionados (Stock / Ajuste / IFL / Gestión), una fila
TOTAL por ajustador seguida de sus filas por tramo, subtotal por división,
total general, y las secciones de abajo (Avance Producción/Meta, Gestión
Comercial, Top 5).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
GRIS = "F2F2F2"
BLANCO = "FFFFFF"

FONT_TITULO = Font(bold=True, size=14, color=BLANCO)
FONT_SUBTITULO = Font(bold=True, size=11, color=BLANCO)
FONT_GRUPO = Font(bold=True, size=9, color=BLANCO)
FONT_LABEL = Font(bold=True, size=9, color=AZUL)
FONT_TOTAL_AJ = Font(bold=True, size=9)
FONT_SUBTOTAL = Font(bold=True, size=9, color=BLANCO)
FONT_NORMAL = Font(size=9)
FONT_SECCION = Font(bold=True, size=12, color=BLANCO)

FILL_TITULO = PatternFill("solid", fgColor=AZUL)
FILL_GRUPO = PatternFill("solid", fgColor=AZUL)
FILL_LABEL = PatternFill("solid", fgColor=AZUL_CLARO)
FILL_TOTAL_AJ = PatternFill("solid", fgColor=GRIS)
FILL_SUBTOTAL = PatternFill("solid", fgColor=AZUL)
FILL_SECCION = PatternFill("solid", fgColor=AZUL)

BORDE_FINO = Border(*(Side(style="thin", color="BFBFBF"),) * 4)

COLUMNAS_GRILLA = [
    ("Division", "División", 22),
    ("Ajustador", "Ajustador", 24),
    ("Asignados", "Asignados", 10),
    ("Stock_Q", "Q", 6),
    ("Dias_prom", "Días prom", 10),
    ("Stock_Hon_UF", "Hon UF", 11),
    ("Ajuste_Qp", "Q p", 6),
    ("Ajuste_Meta", "Meta", 6),
    ("Ajuste_HonpUF", "Hon p UF", 11),
    ("Ajuste_Qr", "Q r", 6),
    ("Ajuste_HonrUF", "Hon r UF", 11),
    ("IFL_Qp", "Q p", 6),
    ("IFL_Meta", "Meta", 6),
    ("IFL_HonpUF", "Hon p UF", 11),
    ("IFL_Qr", "Q r", 6),
    ("IFL_HonrUF", "Hon r UF", 11),
    ("Gestion_Comercial", "Comercial", 32),
    ("Extra", "Extra", 40),
]
N_COLS = len(COLUMNAS_GRILLA)


def _c(ws, fila, col, valor, font=None, fill=None, align=None, num_format=None, border=True):
    celda = ws.cell(row=fila, column=col, value=valor)
    if font:
        celda.font = font
    if fill:
        celda.fill = fill
    celda.alignment = align or Alignment(horizontal="center", vertical="center", wrap_text=True)
    if num_format:
        celda.number_format = num_format
    if border:
        celda.border = BORDE_FINO
    return celda


def _fila_grupo_y_label(ws, fila):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila + 1, end_column=1)
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila + 1, end_column=2)
    _c(ws, fila, 1, "División", font=FONT_LABEL, fill=FILL_LABEL)
    _c(ws, fila, 2, "Ajustador", font=FONT_LABEL, fill=FILL_LABEL)

    grupos = [(3, 6, "Stock"), (7, 11, "Ajuste"), (12, 16, "IFL"), (17, 18, "Gestión")]
    for c_ini, c_fin, nombre in grupos:
        ws.merge_cells(start_row=fila, start_column=c_ini, end_row=fila, end_column=c_fin)
        _c(ws, fila, c_ini, nombre, font=FONT_GRUPO, fill=FILL_GRUPO)
        for c in range(c_ini + 1, c_fin + 1):
            _c(ws, fila, c, None, fill=FILL_GRUPO)

    for idx, (_, etiqueta, _) in enumerate(COLUMNAS_GRILLA):
        if idx < 2:
            continue
        _c(ws, fila + 1, idx + 1, etiqueta, font=FONT_LABEL, fill=FILL_LABEL)
    return fila + 2


def _escribir_grilla_semana(ws, fila, titulo, grilla_con_totales, subtotales):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
    _c(ws, fila, 1, titulo, font=FONT_SUBTITULO, fill=FILL_TITULO, align=Alignment(horizontal="left", vertical="center"))
    fila += 1

    fila = _fila_grupo_y_label(ws, fila)

    if grilla_con_totales.empty:
        _c(ws, fila, 1, "Sin datos para este período.", font=FONT_NORMAL, align=Alignment(horizontal="left"))
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
        return fila + 2

    fila_inicio_division = {}
    ajustador_previo = None
    for _, fila_datos in grilla_con_totales.iterrows():
        division = fila_datos["Division"]
        ajustador = fila_datos["Ajustador"]
        es_primera_fila_ajustador = ajustador != ajustador_previo
        ajustador_previo = ajustador
        if division not in fila_inicio_division:
            fila_inicio_division[division] = fila

        for idx, (clave, _, _) in enumerate(COLUMNAS_GRILLA):
            col_idx = idx + 1
            if clave == "Division":
                valor = division if es_primera_fila_ajustador else None
            elif clave == "Ajustador":
                # Nombre + tramo pegados en la misma fila (igual que el Excel original),
                # nunca el nombre solo en una fila aparte por encima del detalle.
                valor = f"{ajustador} {fila_datos['Tramo']}"
            elif clave in ("Ajuste_Meta", "IFL_Meta"):
                v = fila_datos.get(clave)
                valor = None if v in (None, "") or (isinstance(v, float) and v != v) else v
            else:
                valor = fila_datos.get(clave)

            num_format = "0.00" if "Hon" in clave else ("0" if clave not in ("Division", "Ajustador", "Gestion_Comercial", "Extra") else None)
            align = Alignment(horizontal="left", vertical="center", wrap_text=True) if clave in ("Ajustador", "Gestion_Comercial", "Extra") else None
            _c(ws, fila, col_idx, valor, font=FONT_NORMAL, align=align, num_format=num_format)
        fila += 1

    for division, fila_ini in fila_inicio_division.items():
        fila_fin = fila - 1
        # encuentra el último renglón de esa división (todas sus filas fueron consecutivas)
        idxs = grilla_con_totales.index[grilla_con_totales["Division"] == division]
        fila_fin = fila_ini + len(grilla_con_totales.loc[idxs]) - 1
        if fila_fin > fila_ini:
            ws.merge_cells(start_row=fila_ini, start_column=1, end_row=fila_fin, end_column=1)

    if not subtotales.empty:
        for _, fila_sub in subtotales.iterrows():
            for idx, (clave, _, _) in enumerate(COLUMNAS_GRILLA):
                col_idx = idx + 1
                if clave == "Division":
                    valor = ("Subtotal " + fila_sub["Division"]) if fila_sub["Division"] != "TOTAL GENERAL" else "TOTAL GENERAL"
                elif clave == "Ajustador":
                    valor = None
                elif clave in ("Ajuste_Meta", "IFL_Meta", "Gestion_Comercial", "Extra"):
                    valor = None
                else:
                    valor = fila_sub.get(clave)
                num_format = "0.00" if "Hon" in clave else ("0" if clave not in ("Division", "Ajustador") else None)
                _c(ws, fila, col_idx, valor, font=FONT_SUBTOTAL, fill=FILL_SUBTOTAL, num_format=num_format)
            fila += 1

    return fila + 1


def _titulo_seccion(ws, fila, texto, ancho=N_COLS):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ancho)
    _c(ws, fila, 1, texto, font=FONT_SECCION, fill=FILL_SECCION, align=Alignment(horizontal="left", vertical="center"))
    return fila + 2


def _tabla_simple(ws, fila, encabezados, filas, anchos=None):
    for i, h in enumerate(encabezados):
        _c(ws, fila, i + 1, h, font=FONT_LABEL, fill=FILL_LABEL)
    fila += 1
    if not filas:
        _c(ws, fila, 1, "Sin datos.", font=FONT_NORMAL, align=Alignment(horizontal="left"))
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(len(encabezados), 2))
        return fila + 2
    for fila_datos in filas:
        for i, valor in enumerate(fila_datos):
            align = Alignment(horizontal="left", vertical="center", wrap_text=True) if isinstance(valor, str) else None
            _c(ws, fila, i + 1, valor, font=FONT_NORMAL, align=align)
        fila += 1
    return fila + 1


def generar_workbook_tablero(
    rango_previa, grilla_previa, subtotales_previa,
    rango_actual, grilla_actual, subtotales_actual,
    avance_df, metas_anuales_df,
    bitacora_df,
    top5_ingenieria, top5_movil,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tablero"

    for i, (_, _, ancho) in enumerate(COLUMNAS_GRILLA):
        ws.column_dimensions[get_column_letter(i + 1)].width = ancho

    fila = 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
    _c(ws, fila, 1, "TABLERO — DIVISIÓN INGENIERÍA Y EQUIPO MÓVIL", font=FONT_TITULO, fill=FILL_TITULO,
       align=Alignment(horizontal="left", vertical="center"))
    fila += 2

    fila = _escribir_grilla_semana(ws, fila, f"Semana Anterior  ({rango_previa})", grilla_previa, subtotales_previa)
    fila = _escribir_grilla_semana(ws, fila, f"Semana Actual  ({rango_actual})", grilla_actual, subtotales_actual)

    fila = _titulo_seccion(ws, fila, "AVANCE — PRODUCCIÓN / META")
    filas_avance = []
    if not avance_df.empty:
        metas_dict = {m["division"]: m for m in metas_anuales_df.to_dict(orient="records")} if not metas_anuales_df.empty else {}
        for _, r in avance_df.iterrows():
            meta_info = metas_dict.get(r["Division"], {})
            meta_anual = meta_info.get("meta_anual_uf")
            meta_mes = meta_info.get("meta_mensual_uf")
            pct_anual = f"{round(100 * r['Produccion_Acumulada_UF'] / meta_anual, 1)}%" if meta_anual else "Sin meta"
            pct_mes = f"{round(100 * r['Produccion_Mes_UF'] / meta_mes, 1)}%" if meta_mes else "Sin meta"
            filas_avance.append([r["Division"], meta_anual or "Sin definir", round(r["Produccion_Acumulada_UF"], 2), pct_anual, meta_mes or "Sin definir", round(r["Produccion_Mes_UF"], 2), pct_mes])
    fila = _tabla_simple(ws, fila, ["Área", "Meta Anual (UF)", "Producción Acumulada (UF)", "% Cump Anual", "Meta Mes (UF)", "Producción Mes (UF)", "% Cump Mes"], filas_avance)

    fila = _titulo_seccion(ws, fila, "GESTIÓN COMERCIAL / REUNIONES DE TRABAJO")
    filas_bitacora = bitacora_df[["Ajustador", "Detalle", "Fecha"]].values.tolist() if not bitacora_df.empty else []
    fila = _tabla_simple(ws, fila, ["Ajustador", "Detalle", "Fecha"], filas_bitacora)

    fila = _titulo_seccion(ws, fila, "TOP FIVE INGENIERÍA")
    filas_top = list(zip(top5_ingenieria[0].index, top5_ingenieria[0].values)) if not top5_ingenieria[0].empty else []
    fila = _tabla_simple(ws, fila, ["Compañía de seguros", "Casos vigentes"], filas_top)
    filas_top = list(zip(top5_ingenieria[1].index, top5_ingenieria[1].values)) if not top5_ingenieria[1].empty else []
    fila = _tabla_simple(ws, fila, ["Corredora", "Casos vigentes"], filas_top)

    fila = _titulo_seccion(ws, fila, "TOP FIVE EQUIPO MÓVIL")
    filas_top = list(zip(top5_movil[0].index, top5_movil[0].values)) if not top5_movil[0].empty else []
    fila = _tabla_simple(ws, fila, ["Compañía de seguros", "Casos vigentes"], filas_top)
    filas_top = list(zip(top5_movil[1].index, top5_movil[1].values)) if not top5_movil[1].empty else []
    fila = _tabla_simple(ws, fila, ["Corredora", "Casos vigentes"], filas_top)

    fila = _titulo_seccion(ws, fila, "ESTÁNDARES DE SERVICIO (SLA)")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
    _c(ws, fila, 1,
       "Pendiente: esta sección requiere fechas por hito (contacto, inspección, informe preliminar, rechazo, addendum) "
       "que hoy no se registran en el planificador. Se puede habilitar si se agrega ese registro.",
       font=FONT_NORMAL, align=Alignment(horizontal="left", wrap_text=True))
    ws.row_dimensions[fila].height = 30

    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
